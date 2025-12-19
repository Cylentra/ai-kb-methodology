#!/usr/bin/env python3
"""
Excel to Markdown Converter

Converts .xlsx files to Markdown format, preserving:
- Multiple sheets (as separate sections)
- Tables with headers
- Basic cell formatting

Usage:
    python xlsx_to_markdown.py input.xlsx [output.md]
    python xlsx_to_markdown.py /path/to/folder  # Batch convert all .xlsx files

Requirements:
    pip install openpyxl
"""

import sys
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def clean_cell_value(value):
    """Clean and format a cell value for Markdown."""
    if value is None:
        return ""

    # Convert to string
    text = str(value).strip()

    # Escape pipe characters (used in Markdown tables)
    text = text.replace("|", "\\|")

    # Replace newlines with <br> for Markdown compatibility
    text = text.replace("\n", "<br>")

    return text


def convert_sheet_to_markdown(sheet, sheet_name):
    """Convert a single worksheet to Markdown."""
    md_parts = []

    # Sheet header
    md_parts.append(f"## {sheet_name}")

    # Get all rows with data
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        md_parts.append("*Empty sheet*")
        return "\n\n".join(md_parts)

    # Filter out completely empty rows
    rows = [row for row in rows if any(cell is not None for cell in row)]

    if not rows:
        md_parts.append("*Empty sheet*")
        return "\n\n".join(md_parts)

    # Determine the actual number of columns (ignore trailing empty columns)
    max_cols = 0
    for row in rows:
        for i, cell in enumerate(reversed(row)):
            if cell is not None:
                max_cols = max(max_cols, len(row) - i)
                break

    if max_cols == 0:
        md_parts.append("*Empty sheet*")
        return "\n\n".join(md_parts)

    # Build Markdown table
    table_rows = []

    for row_idx, row in enumerate(rows):
        # Pad row to max_cols if needed
        cells = list(row[:max_cols])
        while len(cells) < max_cols:
            cells.append(None)

        # Clean cell values
        cleaned_cells = [clean_cell_value(cell) for cell in cells]

        # Create table row
        table_row = "| " + " | ".join(cleaned_cells) + " |"
        table_rows.append(table_row)

        # Add header separator after first row
        if row_idx == 0:
            separator = "| " + " | ".join(["---"] * max_cols) + " |"
            table_rows.append(separator)

    md_parts.append("\n".join(table_rows))

    return "\n\n".join(md_parts)


def convert_xlsx_to_markdown(input_path, output_path=None):
    """
    Convert an Excel file to Markdown.

    Args:
        input_path: Path to the .xlsx file
        output_path: Optional output path. If None, uses input filename with .md extension

    Returns:
        Path to the output file
    """
    input_path = Path(input_path)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    if input_path.suffix.lower() not in ['.xlsx', '.xls']:
        raise ValueError(f"Input file must be an Excel file: {input_path}")

    # Determine output path
    if output_path is None:
        output_path = input_path.with_suffix('.md')
    else:
        output_path = Path(output_path)

    # Load workbook
    wb = load_workbook(str(input_path), data_only=True)

    # Build Markdown content
    md_parts = []

    # Document title (from filename)
    doc_title = input_path.stem.replace("_", " ").replace("-", " ").title()
    md_parts.append(f"# {doc_title}")
    md_parts.append(f"*Converted from: {input_path.name}*")

    # Add table of contents if multiple sheets
    if len(wb.sheetnames) > 1:
        md_parts.append("\n## Contents")
        for sheet_name in wb.sheetnames:
            # Create anchor link
            anchor = sheet_name.lower().replace(" ", "-")
            md_parts.append(f"- [{sheet_name}](#{anchor})")

    md_parts.append("---")

    # Convert each sheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_md = convert_sheet_to_markdown(sheet, sheet_name)
        md_parts.append(sheet_md)
        md_parts.append("---")

    # Write output
    markdown_content = "\n\n".join(md_parts)
    output_path.write_text(markdown_content, encoding='utf-8')

    print(f"Converted: {input_path.name} -> {output_path.name} ({len(wb.sheetnames)} sheets)")
    return output_path


def batch_convert(folder_path):
    """Convert all .xlsx files in a folder."""
    folder = Path(folder_path)

    if not folder.is_dir():
        raise NotADirectoryError(f"Not a directory: {folder}")

    xlsx_files = list(folder.glob("*.xlsx")) + list(folder.glob("*.xls"))

    if not xlsx_files:
        print(f"No Excel files found in {folder}")
        return []

    converted = []
    for xlsx_file in xlsx_files:
        try:
            output_path = convert_xlsx_to_markdown(xlsx_file)
            converted.append(output_path)
        except Exception as e:
            print(f"Error converting {xlsx_file.name}: {e}")

    print(f"\nConverted {len(converted)} of {len(xlsx_files)} files")
    return converted


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    input_path = Path(sys.argv[1])

    # Check if input is a directory (batch mode) or single file
    if input_path.is_dir():
        batch_convert(input_path)
    else:
        output_path = sys.argv[2] if len(sys.argv) > 2 else None
        convert_xlsx_to_markdown(input_path, output_path)


if __name__ == "__main__":
    main()
